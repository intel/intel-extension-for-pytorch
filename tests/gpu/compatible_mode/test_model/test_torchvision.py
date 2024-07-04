# runn all models usage:
# under the torchbench folder and python -u test_torchbench.py
# run single model usage:
# bs is 0 means use the default value
# python -u run.py -d xpu -t train --bs 0 --metrics None Model_name
# bs is 4 means set the bs to 4
# python -u run.py -d xpu -t train --bs 4 --metrics None Model_name

import os
import subprocess
from typing import List, Dict, Any
import sys
import time
import json
from torchvision.models import list_models
from torchvision import models


# config
run_bs = 4 # default 4, it is used the batch size which model can be set
executor = sys.executable
device = 'cuda' # default xpu
unsupported_str_list = ["train", "notimplement", "support"]
unsupported_amp_str = "error: unrecognized arguments: --amp"
onlysupportcpu_str = "The eval test only supports CPU"
#samewithcuda_str = ["expected scalar type Float but found Half", "mat1 and mat2 must have the same dtype, but got Float and Half", "Unsupported input type: <class", "attempting to assign a gradient with dtype", "Found dtype Float but expected Half", "\"host_softmax\" not implemented for \'Long\'", "Input type (float) and bias type (c10::Half) should be the same", "The instance variable \'model\' does not exist or is not type \'torch.nn.Module\'", "ValueError: input must have the type torch.float16, got type torch.float32"]
samewithcuda = {
        "trainfp32eager_result": ["cm3leon_generate"],
        "trainfp16eager_result": [
            "DALLE2_pytorch",
            "LearningToPaint",
            "cm3leon_generate",
            "demucs",
            "detectron2_maskrcnn",
            "dlrm",
            "drq",
            "fastNLP_Bert",
            "functorch_dp_cifar10",
            "lennard_jones",
            "maml",
            "maml_omniglot",
            "mobilenet_v2_quantized_qat",
            "moco",
            "nvidia_deeprecommender",
            "opacus_cifar10",
            "pytorch_stargan",
            "resnet50_quantized_qat",
            "soft_actor_critic",
            "speech_transformer",
            "tacotron2",
            "tts_angular",
            "vision_maskrcnn",],
        "inffp32eager_result": ["Background_Matting"],
        "inffp16eager_result": [
            "DALLE2_pytorch",
            "drq",
            "fastNLP_Bert",
            "lennard_jones",
            "maml",
            "moco",
            "nvidia_deeprecommender",
            "pytorch_CycleGAN_and_pix2pix",
            "soft_actor_critic",
            "speech_transformer",
            "tts_angular",
            "vision_maskrcnn",
            "Background_Matting",
            ],
        }
mode = os.getenv('TEST_TORCHVISION_MODE')
# default is testing train/inf
if mode is None:
    mode = 'all'
assert type(mode) is str and mode.lower() in ['train', 'inf', 'all'], "please specify the TEST_TORCHVISION_MODE to be train/inf/all"
bench_file = 'train.py'
config_arg = '-u' # python -u xxxx
multi_card = False # default single card running

# log print
MODEL_BLANK_SPACE = 50
CMD_BLANK_SPACE = 150

# all models scope in torchvision
all_models_dict = {
        "classification": list_models(models),
        "detection": list_models(models.detection),
        "segmentation": list_models(models.segmentation),
        "video": list_models(models.video),
        "optical_flow": list_models(models.optical_flow),
        "quantization": list_models(models.quantization),
        #"detection": ["fasterrcnn_mobilenet_v3_large_320_fpn"],
        #"classification": ["resnet50"],
        }
for key in all_models_dict.keys():
    print(key)
common_args_dict = {
                "classification": {
                    "lr": "0.1",
                    "datapath": "imagenet"
                    },
                "detection": {
                    "lr": "0.02",
                    "datapath" : "coco"
                    },
                "segmentation": {
                    "lr": "0.02",
                    "datapath" : "coco",
                    },
                "video": {
                    "datapath": "tiny-Kinetics-400",
                    "lr": "0.02",
                    },
                "optical_flow": {
                    "lr": "0.0004",
                    },
                "quantization": {
                    "lr": "0.1",
                    "datapath": "imagenet",
                    },
                }
extra_category_args = {
                "classification": {
                    },
                "detection": {
                    },
                "segmentation": {
                    },
                "video": {
                    "inf": ["--kinetics-version", "400", "--batch-size", "2", "--clip-len", "1", "--frame-rate", "1"],
                    "train": ["--kinetics-version", "400", "--batch-size", "2", "--clip-len", "1", "--frame-rate", "1"],
                    #"inf": "--kinetics-version 400 --batch-size 2 --clip-len 1 --frame-rate 1",
                    #"train": "--kinetics-version 400 --batch-size 2 --clip-len 1 --frame-rate 1",
                    },
                "optical_flow": {
                    "inf": ["--val-dataset", "kitti", "--batch-size", "1", "--dataset-root", "."],
                    "train": ["--train-dataset", "kitti", "--val-dataset", "kitti", "--batch-size", "2", "--dataset-root", "."],
                    },
                "quantization": {
                    "inf": ["--qbackend", "fbgemm", "--batch-size", "1",],
                    "train": ["--post-training-quantize", "--qbackend", "fbgemm", "--batch-size", "1"],
                    },
                }
extra_args = {
        "ssd300_vgg16": {
            "train": ["--batch-size", "4"],
            }
        }

#all_models_list = [
#            "BERT_pytorch"]

# This list contains the model which is unsupported to set the maunal bs
# so set the bs 0 in command
models_list_unsupport_manual_bs = [
    "basic_gnn_edgecnn",
    "basic_gnn_gcn",
    "basic_gnn_gin",
    "basic_gnn_sage",
    "Background_Matting",
    "functorch_maml_omniglot",
    "maml",
    "maml_omniglot",
    "pytorch_CycleGAN_and_pix2pix",
    "pytorch_stargan",
    "soft_actor_critic",
    "speech_transformer",
    "stable_diffusion_text_encoder",
    "stable_diffusion_unet",
    "torch_multimodal_clip",
    "vision_maskrcnn",
    "pyhpc_turbulent_kinetic_energy",
    "drq",
]

# This list contains the model which is unsupported to train
# the train function is not implemented by torchbench
models_list_unsupport_eager_mode = [
    "simple_gpt",
    "simple_gpt_tp_manual",
]

# This list contains the model which needs to set/unset the specific env flag
# "model": {"set": {"env_name":"env_value"}, "unset":["env_name"]}
models_list_specific_env_flag = {
    "moco":            {"set": {"CCL_ZE_IPC_EXCHANGE": "sockets"}, "unset": ["ZE_AFFINITY_MASK"]},
    "vision_maskrcnn": {"set": {"CCL_ZE_IPC_EXCHANGE": "sockets"}, "unset": ["ZE_AFFINITY_MASK"]},
}

# contain the required bs for specific models
specific_models_with_bs: Dict[str, int] = {
    "torch_multimodal_clip": 1,
}

# this list contains the precision we will try to test
#precision_list = ["fp32", "fp16", "amp", "bf16","amp_fp16", "amp_bf16"]
#precision_list = ["amp", "amp_fp16"]
if os.getenv("PRECISIONS"):
    precision_list = os.getenv("PRECISIONS").split("|")
else:
    precision_list = ["fp32", "fp16", "bf16",  "amp"]
print(precision_list)
#precision_list = ["amp"]

# this list contains the backends we will try to test
#backend_list = ["eager", "torchscript"]
backend_list = ["eager"]




# example: {"BERT_pytorch": {"bs": 4, "train": True or None, "inf": True or None}}
models_list: Dict[str, Dict[str, Any]] = {}

pass_model_cnt = 0
fail_model_cnt = 0
not_supported_model_cnt = 0
# example: {"BERT_pytorch": {"result_key": {"pass"  : True}}}
#          {"BERT_pytorch": {"result_key": {"failed": "error message"}}}
#          {"BERT_pytorch": {"result_key": {"Notsupported": "not supported message"}}}
#          {"BERT_pytorch": {"duration": int}}
results_summary: Dict[str, Dict[str, Any]] = {}

def is_eligible_for_test(mode, model):
    if model in models_list_unsupport_eager_mode:
        return False
    return True

def set_the_running_mode(model):
    if mode == 'train':
        models_list[model]['train'] = True
        models_list[model]['inf'] = False
    elif mode == 'inf':
        models_list[model]['train'] = False
        models_list[model]['inf'] = True
    elif mode == 'all':
        models_list[model]['train'] = True
        models_list[model]['inf'] = True
    else:
        raise RuntimeError("[error] no mode is specified, please check env flag TEST_TORCHBENCH_MODE")

def set_the_batch_size(model):
    if model in models_list_unsupport_manual_bs:
        models_list[model]['bs'] = 0
    elif model in specific_models_with_bs:
        models_list[model]['bs'] = specific_models_with_bs[model]
    else:
        models_list[model]['bs'] = run_bs

for category, model_list in all_models_dict.items():
    # avoid the recursive deal with model
    for model in model_list:
        if model in models_list:
            continue
        # skip the model which unsupport train/inf in torchbench
        if not is_eligible_for_test(mode, model):
            continue

        models_list[model] = {"category": category}

        set_the_running_mode(model)
        set_the_batch_size(model)

# models are eligible to run
num_models = len(models_list.keys())
assert num_models >= 1, "[error] no model is collected"
print('[info] using python executor: ', executor)
print('[info] mode: ', mode)
print('[info] device: ', device)
print('[info] bs: ', run_bs)
print('[info] multi card: ', multi_card)
print('[info] running models list:')
for idx, elem in enumerate(models_list.items()):
    print('[', idx, '] ', elem)

def generate_required_env(model):
    # set the running env
    env = os.environ.copy()

    # remove the env IPEX_SHOW_OPTION for better logging
    env['IPEX_SHOW_OPTION'] = f"0"

    # single card running default
    if not multi_card:
        env['ZE_AFFINITY_MASK'] = f"0"

    if model in models_list_specific_env_flag:
        required_env = models_list_specific_env_flag[model]
        assert "set" in required_env and "unset" in required_env, "check the models_list_specific_env_flag, it is wrongly set"

        # set the required env flag
        for env_name, env_value in required_env["set"].items():
            env[env_name] = env_value

        # remove the env flag
        for env_name in required_env["unset"]:
            del env[env_name]
    return env

def generate_command(model, category, test_mode='train', precision='fp32', backend='eager'):
    cmd: List[str] = []
    global bench_file
    # generate the cmd
    # python -u references/detection/train.py --lr 0.02 --epochs  1  --device cpu --data-path /home/gta/wliao2/torchvision/coco --model fasterrcnn_mobilenet_v3_large_320_fpn --aspect-ratio-group-factor 3 --weights-backbone MobileNet_V3_Large_Weights.IMAGENET1K_V1
    cmd.append(str(executor))
    cmd.append(config_arg)
    if category == 'quantization':
        bench_path = "vision/references/classification/"
        bench_file = "train_quantization.py"
    elif category == 'video':
        bench_path = "vision/references/video_classification/"
    else:        
        bench_path = "vision/references/" + category + "/"
    cmd.append(bench_path + bench_file)
    print(f'cmd is {cmd}')
    # specify the model
    cmd.append('--model')
    cmd.append(model)
    cmd.append('--lr')
    cmd.append(common_args_dict[category]["lr"])
    if test_mode == 'inf' and models_list[model]['inf'] and category != "optical_flow":
        cmd.append('--test-only')
    cmd.append('--epochs')
    cmd.append('1')
    # disable metrics because it will init pynvml
    cmd.append('--device')
    cmd.append(device)
    if 'keypoint' in model:
        cmd.append('--dataset')
        cmd.append('coco_kp')
        cmd.append('--data-path')
        cmd.append('coco_kp')
        cmd.append('--batch-size')
        cmd.append('1')
    elif category != "optical_flow":
        cmd.append('--data-path')
        cmd.append(common_args_dict[category]['datapath'])
    if precision == "amp":
        cmd.append('--amp')
    elif precision != "default":
        cmd.append('--precision')
        cmd.append(precision)
    #if backend != "eager":
    #    cmd.append('--backend')
    #    cmd.append(backend)
    if extra_category_args[category]:
        if test_mode in extra_category_args[category]:
            cmd.extend(extra_category_args[category][test_mode])
    if model in extra_args.keys() and test_mode in extra_args[model].keys():
        cmd.extend(extra_args[model][test_mode])
    # specify the model
    #cmd.append(model)
    str_cmd = ''
    print(cmd)
    for elem in cmd:
        str_cmd += elem + ' '
    return cmd, str_cmd.rstrip()

def testing(model, category, test_mode, precison, backend, count):
    global pass_model_cnt
    global fail_model_cnt
    global not_supported_model_cnt
    result_key = test_mode + precision + backend + '_result'
    if model not in results_summary:
        results_summary[model] = {}
    if result_key not in results_summary[model]:
        results_summary[model][result_key] = {}
    print(f'{category},{model}')
    cmd, str_cmd = generate_command(model, category, test_mode, precision, backend)
    env = generate_required_env(model)
    stderr_msg = None
    returncode = 0
    is_not_supported = False
    time_start = time.time()
    try:
        # start run the model
        print(f'begin to run cmd" {cmd} {str_cmd}')
        subprocess.run(cmd, encoding="utf-8", capture_output=True, check=True, env=env)
        print("finished")
    except subprocess.CalledProcessError as e:
        returncode = e.returncode
        if returncode != 0:
            stderr_msg = e.stderr
            # check if it is not implemented
            cond1 = unsupported_amp_str in stderr_msg
            cond2 = onlysupportcpu_str in stderr_msg
            cond3 = model in samewithcuda[result_key] if result_key in samewithcuda.keys() else False
            is_not_supported = all([key.lower() in stderr_msg.lower() for key in unsupported_str_list]) or cond1 or cond2 or cond3
    time_end = time.time()
    # set the duration
    duration = round(time_end - time_start, 2)
    results_summary[model]['duration'] = duration

    # for log alignment
    model_space = MODEL_BLANK_SPACE
    model_space = ' ' * (model_space - len(model))
    cmd_space = CMD_BLANK_SPACE
    cmd_space = ' ' * (cmd_space - len(str_cmd))
    if test_mode == 'inf':
        test_mode += ' ' * 2
    results_summary[model][result_key]['category'] = category

    # pass
    if returncode == 0:
        print('[', count, '][', test_mode, '][success] pass model: ', model, model_space, 'cmd: ', str_cmd, cmd_space, 'time(s): ', duration)
        pass_model_cnt += 1
        results_summary[model][result_key]['pass'] = True
    else:
        print('[', count, '][', test_mode, '][error]   fail model: ', model, model_space, 'cmd: ', str_cmd, cmd_space, 'time(s): ', duration)
        print(stderr_msg.splitlines()[-1])
        # not supported train or inf
        if is_not_supported:
            #print(f'-----------{model},{result_key}--------')
            not_supported_model_cnt += 1
            results_summary[model][result_key]['Notsupported'] = stderr_msg
            results_summary[model][result_key]['simple_Notsupported'] = stderr_msg.splitlines()[-1]
        else:
            # real failed
            fail_model_cnt += 1
            results_summary[model][result_key]['failed'] = stderr_msg
            results_summary[model][result_key]['simple_failed'] = stderr_msg.splitlines()[-1]

print('[info] running models number: ', num_models)
print('[info] begin test all models......')
global_time_start = time.time()
count = 0
for model in models_list.keys():
    for precision in precision_list:
        for backend in backend_list:
            if models_list[model]['train']:
                testing(model, models_list[model]['category'], 'train', precision, backend, count)
                count += 1
            if models_list[model]['inf']:
                testing(model, models_list[model]['category'], 'inf', precision, backend, count)
                count += 1

global_time_end = time.time()
global_duration = global_time_end - global_time_start

# summary
print('\n' * 10 + '*' * 50)
print('[info] Detailed Summary:')
for idx, model in enumerate(results_summary.keys()):
    assert model in results_summary, "Fatal error, the model is not in the testing records"

    model_space = MODEL_BLANK_SPACE
    model_space = ' ' * (model_space - len(model))
    for test_mode in ['train', 'inf']:
        for precision in precision_list:
            for backend in backend_list:
                result_key = test_mode + precision + backend + '_result'
                if result_key in results_summary[model]:
                    if 'pass' in results_summary[model][result_key]:
                        print('[', idx, '] model ', model, model_space, test_mode, precision, backend, ' pass')
                    elif 'failed' in results_summary[model][result_key]:
                        print('[', idx, '] model ', model, model_space, test_mode, precision, backend, ' failed:\n', results_summary[model][result_key]['failed'])
                    elif 'Notsupported' in results_summary[model][result_key]:
                        print('[', idx, '] model ', model, model_space, test_mode, precision, backend, ' not supported:\n', results_summary[model][result_key]['Notsupported'])
                    else:
                        raise RuntimeError("model {} is not recorded into the results, check if it is running".format(model))

print('\n' * 10 + '*' * 50)
print('[info] Simplified Summary:')
with open("results_summary.log", "w") as f:
    json.dump(results_summary, f)

summary = {}
details = {}
#import pdb;pdb.set_trace()
for category in all_models_dict.keys():
    summary[category] = {"error": 0, "passes": 0, "notsupport": 0, "total": 0}
    for test_mode in ['train', 'inf']:
        summary[category + "_" + test_mode] = {"error": 0, "passes": 0, "notsupport": 0, "total": 0}
        for precision in precision_list:
            summary[category + "_" + test_mode + "_" + precision] = {"error": 0, "passes": 0, "notsupport": 0, "total": 0}
            for backend in backend_list:
                summary[category + "_" + test_mode + "_" + precision + "_" + backend] = {"error": 0, "passes": 0, "notsupport": 0, "total": 0}
                details[category + "_" + test_mode + "_" + precision + "_" + backend] = {"error": [], "passes": [], "notsupport": []}
for idx, model in enumerate(results_summary.keys()):
    for test_mode in ['train', 'inf']:
        for precision in precision_list:
            for backend in backend_list:
                result_key = test_mode + precision + backend + '_result'
                category = models_list[model]['category']
            if result_key in results_summary[model]:
                if 'pass' in results_summary[model][result_key]:
                    print('[', idx, '] model ', category, model, test_mode, precision, backend, ' pass')
                    summary[category + "_" + test_mode + "_" + precision + "_" + backend]["passes"] += 1
                    summary[category + "_" + test_mode + "_" + precision + "_" + backend]["total"] += 1
                    summary[category + "_" + test_mode + "_" + precision ]["passes"] += 1
                    summary[category + "_" + test_mode + "_" + precision ]["total"] += 1
                    summary[category + "_" + test_mode ]["passes"] += 1
                    summary[category + "_" + test_mode ]["total"] += 1
                    summary[category]["passes"] += 1
                    summary[category]["total"] += 1
                    details[category + "_" + test_mode + "_" + precision + "_" + backend]["passes"].append(model)
                elif 'failed' in results_summary[model][result_key]:
                    print('[', idx, '] model ', category, model, test_mode, precision, backend, ' fail')
                    summary[category + "_" + test_mode + "_" + precision + "_" + backend]["error"] += 1
                    summary[category + "_" + test_mode + "_" + precision + "_" + backend]["total"] += 1
                    summary[category + "_" + test_mode + "_" + precision ]["error"] += 1
                    summary[category + "_" + test_mode + "_" + precision ]["total"] += 1
                    summary[category + "_" + test_mode ]["error"] += 1
                    summary[category + "_" + test_mode ]["total"] += 1
                    summary[category]["error"] += 1
                    summary[category]["total"] += 1
                    details[category + "_" + test_mode + "_" + precision + "_" + backend]["error"].append(model)
                elif 'Notsupported' in results_summary[model][result_key]:
                    print('[', idx, '] model ', category, model, test_mode, precision, backend, ' not supported')
                    summary[category + "_" + test_mode + "_" + precision + "_" + backend]["notsupport"] += 1
                    summary[category + "_" + test_mode + "_" + precision + "_" + backend]["total"] += 1
                    summary[category + "_" + test_mode + "_" + precision ]["notsupport"] += 1
                    summary[category + "_" + test_mode + "_" + precision ]["total"] += 1
                    summary[category + "_" + test_mode ]["notsupport"] += 1
                    summary[category + "_" + test_mode ]["total"] += 1
                    summary[category]["notsupport"] += 1
                    summary[category]["total"] += 1
                    details[category + "_" + test_mode + "_" + precision + "_" + backend]["notsupport"].append(model)
                else:
                    raise RuntimeError("model {} is not recorded with the results, check if it run or not".format(model))

# [watch] calculate the pass rate not includes the not supported model
if pass_model_cnt == 0 and fail_model_cnt == 0:
    print('[Error] No pass models or failed models are found')
    sys.exit()

for key in summary.keys():
    if summary[key]["total"] != 0 and summary[key]["total"] - summary[key]["notsupport"] != 0:
        pass_rate = round(summary[key]["passes"] / summary[key]["total"], 3)
        fail_rate = round(summary[key]["error"] / summary[key]["total"], 3)
        pass_rate_wo = round(summary[key]["passes"] / (summary[key]["total"] - summary[key]["notsupport"]), 3)
    elif summary[key]["total"] == 0:
        pass_rate = 0
        fail_rate = 0
        pass_rate_wo = 0
    elif summary[key]["total"] - summary[key]["notsupport"] == 0:
        pass_rate = 0
        fail_rate = 0
        pass_rate_wo = 1
    summary[key]["passrate"] = pass_rate
    summary[key]["failrate"] = fail_rate
    summary[key]["passrate_wo"] = pass_rate_wo
    print(f'[info] {key} pass number = {summary[key]["passes"]}')
    print(f'[info] {key} fail number = {summary[key]["error"]}')
    print(f'[info] {key} not support number = {summary[key]["notsupport"]}')
    print(f'[info] {key} total number = {summary[key]["total"]}')
    print(f'[info] {key} pass rate without not supported= {pass_rate_wo}')
    print(f'[info] {key} pass rate = {pass_rate}')
    print(f'[info] {key} fail rate = {fail_rate}')

print('[info] eligible testing model number = ', num_models)
global_duration = global_duration / 60.0 # unit: min
print('[info] testing total duration = ', round(global_duration, 2), ' min')

with open("summary.log", 'w') as f:
    json.dump(summary, f)
with open("details.log", 'w') as f:
    json.dump(details, f)

from openpyxl import Workbook
wb = Workbook()
#sheet = wb.active

summary_total = {}
#sheet.append(("modelname", "usecase", "backend", "precision", "result")
ws_s = wb.create_sheet(title="summary")
ws_s.append(("type", "pass", "fail", "notsupport", "total", "passrate w/o unsupported", "passrate", "failrate"))
for key in summary:
    if "_eager" in key:
        continue
    print((key, {summary[key]["passes"]}, {summary[key]["error"]}, {summary[key]["notsupport"]}, {summary[key]["total"]}, summary[key]["passrate_wo"], summary[key]["passrate"], summary[key]["failrate"]))
    ws_s.append((key, summary[key]["passes"], summary[key]["error"], summary[key]["notsupport"], summary[key]["total"], summary[key]["passrate_wo"],  summary[key]["passrate"], summary[key]["failrate"]))
    if "train_" in key or "inf_" in key:
        key1 = key.split("_train_")[1] if "train" in key else key.split("_inf_")[1]
        if key1 not in summary_total.keys():
            if summary[key]["total"] - summary[key]["notsupport"] != 0:
                passrate_wo = round(summary[key]["passes"]/(summary[key]["total"] - summary[key]["notsupport"]), 3)
            else:
                passrate_wo = -1
            passrate = round(summary[key]["passes"]/summary[key]["total"], 3)
            failrate = round(summary[key]["error"]/summary[key]["total"], 3)
            summary_total[key1] = {"passes": summary[key]["passes"], "error": summary[key]["error"], "notsupport": summary[key]["notsupport"], "total": summary[key]["total"], "passrate_wo": passrate_wo, "passrate": passrate, "failrate": failrate}
        else:
            summary_total[key1]["passes"] += summary[key]["passes"]
            summary_total[key1]["error"] += summary[key]["error"]
            summary_total[key1]["notsupport"] += summary[key]["notsupport"]
            summary_total[key1]["total"] += summary[key]["total"]
            if summary[key]["total"] - summary[key]["notsupport"] != 0:
                passrate_wo = round(summary_total[key1]["passes"]/(summary_total[key1]["total"] - summary_total[key1]["notsupport"]), 3)
            else:
                passrate_wo = -1
            passrate = round(summary_total[key1]["passes"]/summary_total[key1]["total"], 3)
            failrate = round(summary_total[key1]["error"]/summary_total[key1]["total"], 3)
            summary_total[key1] = {"passes": summary_total[key1]["passes"], "error": summary_total[key1]["error"], "notsupport": summary_total[key1]["notsupport"], "total": summary_total[key1]["total"], "passrate_wo": passrate_wo, "passrate": passrate, "failrate": failrate}
    elif "train" not in key and "inf" not in key:
        key1 = "total"
        if key1 not in summary_total.keys():
            if summary[key]["total"] - summary[key]["notsupport"] != 0:
                passrate_wo = round(summary[key]["passes"]/(summary[key]["total"] - summary[key]["notsupport"]), 3)
            else:
                passrate_wo = -1
            passrate = round(summary[key]["passes"]/summary[key]["total"], 3)
            failrate = round(summary[key]["error"]/summary[key]["total"], 3)
            summary_total[key1] = {"passes": summary[key]["passes"], "error": summary[key]["error"], "notsupport": summary[key]["notsupport"], "total": summary[key]["total"], "passrate_wo": passrate_wo, "passrate": passrate, "failrate": failrate}
        else:
            summary_total[key1]["passes"] += summary[key]["passes"]
            summary_total[key1]["error"] += summary[key]["error"]
            summary_total[key1]["notsupport"] += summary[key]["notsupport"]
            summary_total[key1]["total"] += summary[key]["total"]
            if summary[key]["total"] - summary[key]["notsupport"] != 0:
                passrate_wo = round(summary_total[key1]["passes"]/(summary_total[key1]["total"] - summary_total[key1]["notsupport"]), 3)
            else:
                passrate_wo = -1
            passrate = round(summary_total[key1]["passes"]/summary_total[key1]["total"], 3)
            failrate = round(summary_total[key1]["error"]/summary_total[key1]["total"], 3)
            summary_total[key1] = {"passes": summary_total[key1]["passes"], "error": summary_total[key1]["error"], "notsupport": summary_total[key1]["notsupport"], "total": summary_total[key1]["total"], "passrate_wo": passrate_wo, "passrate": passrate, "failrate": failrate}


for key in summary_total:
    print((key, {summary_total[key]["passes"]}, {summary_total[key]["error"]}, {summary_total[key]["notsupport"]}, {summary_total[key]["total"]}, summary_total[key]["passrate_wo"], summary_total[key]["passrate"], summary_total[key]["failrate"]))
    ws_s.append((key, summary_total[key]["passes"], summary_total[key]["error"], summary_total[key]["notsupport"], summary_total[key]["total"], summary_total[key]["passrate_wo"], summary_total[key]["passrate"], summary_total[key]["failrate"]))


for key in details:
    ws = wb.create_sheet(title=key)
    ws.append(("modelname", "usecase", "backend", "precision", "result", "simple detail", "result details"))
    index = 0
    if "optical_flow" in key:
        category = "optical_flow"
        index = 1
    else:
        category = key.split("_")[0]
    usecase = key.split("_")[index + 1]
    if "amp_fp16" in key or "amp_bf16" in key:
        precision = key.split("_")[index + 2] + "_" + key.split("_")[index + 3]
        backend = key.split("_")[index + 4]
    else:
        precision = key.split("_")[index + 2]
        backend = key.split("_")[index + 3]
    for result, detail in details[key].items():
        for mymodel in detail:
            result_key = usecase + precision + backend + "_result"
            if result == "error":
                result_detail = results_summary[mymodel][result_key]['failed']
                simple_detail = results_summary[mymodel][result_key]['simple_failed']
            elif result == "notsupport":
                result_detail = results_summary[mymodel][result_key]['Notsupported']
                simple_detail = results_summary[mymodel][result_key]['simple_Notsupported']
            else:
                result_detail = ""
                simple_detail = ""
            ws.append((mymodel, usecase, backend, precision, result, simple_detail, result_detail))

wb.save("torchvision.xlsx")



