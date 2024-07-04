# runn all models usage:
# under the torchbench folder and python -u test_torchbench.py
# run single model usage:
# bs is 0 means use the default value
# python -u run.py -d xpu -t train --bs 0 --metrics None Model_name
# bs is 4 means set the bs to 4
# python -u run.py -d xpu -t train --bs 4 --metrics None Model_name

import os
import subprocess
from typing import List, Dict, Any
import sys
import time
import json

# config
run_bs = 4 # default 4, it is used the batch size which model can be set
executor = sys.executable
device = 'cuda' # default xpu
unsupported_str_list = ["train", "notimplement", "support"]
unsupported_amp_str = "doesn't have amp_context support"
onlysupportcpu_str = "The eval test only supports CPU"
#samewithcuda_str = ["expected scalar type Float but found Half", "mat1 and mat2 must have the same dtype, but got Float and Half", "Unsupported input type: <class", "attempting to assign a gradient with dtype", "Found dtype Float but expected Half", "\"host_softmax\" not implemented for \'Long\'", "Input type (float) and bias type (c10::Half) should be the same", "The instance variable \'model\' does not exist or is not type \'torch.nn.Module\'", "ValueError: input must have the type torch.float16, got type torch.float32"]
samewithcuda = {
        "trainfp32eager_result": ["cm3leon_generate"],
        "trainfp16eager_result": [
            "DALLE2_pytorch",
            "LearningToPaint",
            "cm3leon_generate",
            "demucs",
            "detectron2_maskrcnn",
            "dlrm",
            "drq",
            "fastNLP_Bert",
            "functorch_dp_cifar10",
            "lennard_jones",
            "maml",
            "maml_omniglot",
            "mobilenet_v2_quantized_qat",
            "moco",
            "nvidia_deeprecommender",
            "opacus_cifar10",
            "pytorch_stargan",
            "resnet50_quantized_qat",
            "soft_actor_critic",
            "speech_transformer",
            "tacotron2",
            "tts_angular",
            "vision_maskrcnn",],
        "inffp32eager_result": ["Background_Matting"],
        "inffp16eager_result": [
            "DALLE2_pytorch",
            "drq",
            "fastNLP_Bert",
            "lennard_jones",
            "maml",
            "moco",
            "nvidia_deeprecommender",
            "pytorch_CycleGAN_and_pix2pix",
            "soft_actor_critic",
            "speech_transformer",
            "tts_angular",
            "vision_maskrcnn",
            "Background_Matting",
            ],
        }
mode = os.getenv('TEST_TORCHBENCH_MODE')
# default is testing train/inf
if mode is None:
    mode = 'all'
assert type(mode) is str and mode.lower() in ['train', 'inf', 'all'], "please specify the TEST_TORCHBENCH_MODE to be train/inf/all"
bench_file = 'run.py'
config_arg = '-u' # python -u xxxx
multi_card = False # default single card running

# log print
MODEL_BLANK_SPACE = 50
CMD_BLANK_SPACE = 150

# all models scope in torchbench
# TODO: maybe need some methods to get all models enabled in torchbench
all_models_list = [
    "BERT_pytorch",
    "Background_Matting",
    "DALLE2_pytorch",
    "LearningToPaint",
    "Super_SloMo",
    "alexnet",
    "basic_gnn_edgecnn",
    "basic_gnn_gcn",
    "basic_gnn_gin",
    "basic_gnn_sage",
    "cm3leon_generate",
    "dcgan",
    "demucs",
    "densenet121",
    "detectron2_fasterrcnn_r_101_c4",
    "detectron2_fasterrcnn_r_101_dc5",
    "detectron2_fasterrcnn_r_101_fpn",
    "detectron2_fasterrcnn_r_50_c4",
    "detectron2_fasterrcnn_r_50_dc5",
    "detectron2_fasterrcnn_r_50_fpn",
    "detectron2_fcos_r_50_fpn",
    "detectron2_maskrcnn",
    "detectron2_maskrcnn_r_101_c4",
    "detectron2_maskrcnn_r_101_fpn",
    "detectron2_maskrcnn_r_50_c4",
    "detectron2_maskrcnn_r_50_fpn",
    "dlrm",
    "doctr_det_predictor",
    "doctr_reco_predictor",
    "drq",
    "fastNLP_Bert",
    "functorch_dp_cifar10",
    "functorch_maml_omniglot",
    "hf_Albert",
    "hf_Bart",
    "hf_Bert",
    "hf_Bert_large",
    "hf_BigBird",
    "hf_DistilBert",
    "hf_GPT2",
    "hf_GPT2_large",
    "hf_Longformer",
    "hf_Reformer",
    "hf_T5",
    "hf_T5_base",
    "hf_T5_generate",
    "hf_T5_large",
    "hf_Whisper",
    "hf_clip",
    "hf_distil_whisper",
    "lennard_jones",
    "llama",
    "llama_v2_7b_16h",
    "maml",
    "maml_omniglot",
    "mnasnet1_0",
    "mobilenet_v2",
    "mobilenet_v2_quantized_qat",
    "mobilenet_v3_large",
    "moco",
    "nanogpt",
    "nvidia_deeprecommender",
    "opacus_cifar10",
    "phlippe_densenet",
    "phlippe_resnet",
    "pyhpc_equation_of_state",
    "pyhpc_isoneutral_mixing",
    "pyhpc_turbulent_kinetic_energy",
    "pytorch_CycleGAN_and_pix2pix",
    "pytorch_stargan",
    "pytorch_unet",
    "resnet152",
    "resnet18",
    "resnet50",
    "resnet50_quantized_qat",
    "resnext50_32x4d",
    "sam",
    "shufflenet_v2_x1_0",
    "simple_gpt",
    "simple_gpt_tp_manual",
    "soft_actor_critic",
    "speech_transformer",
    "squeezenet1_1",
    "stable_diffusion_text_encoder",
    "stable_diffusion_unet",
    "tacotron2",
    "timm_efficientdet",
    "timm_efficientnet",
    "timm_nfnet",
    "timm_regnet",
    "timm_resnest",
    "timm_vision_transformer",
    "timm_vision_transformer_large",
    "timm_vovnet",
    "torch_multimodal_clip",
    "tts_angular",
    "vgg16",
    "vision_maskrcnn",
    "yolov3"
]
#all_models_list = [
#            "BERT_pytorch"]

# This list contains the model which is unsupported to set the maunal bs
# so set the bs 0 in command
models_list_unsupport_manual_bs = [
    "basic_gnn_edgecnn",
    "basic_gnn_gcn",
    "basic_gnn_gin",
    "basic_gnn_sage",
    "Background_Matting",
    "functorch_maml_omniglot",
    "maml",
    "maml_omniglot",
    "pytorch_CycleGAN_and_pix2pix",
    "pytorch_stargan",
    "soft_actor_critic",
    "speech_transformer",
    "stable_diffusion_text_encoder",
    "stable_diffusion_unet",
    "torch_multimodal_clip",
    "vision_maskrcnn",
    "pyhpc_turbulent_kinetic_energy",
    "drq",
]

# This list contains the model which is unsupported to train
# the train function is not implemented by torchbench
models_list_unsupport_eager_mode = [
    "simple_gpt",
    "simple_gpt_tp_manual",
]

# This list contains the model which needs to set/unset the specific env flag
# "model": {"set": {"env_name":"env_value"}, "unset":["env_name"]}
models_list_specific_env_flag = {
    "moco":            {"set": {"CCL_ZE_IPC_EXCHANGE": "sockets"}, "unset": ["ZE_AFFINITY_MASK"]},
    "vision_maskrcnn": {"set": {"CCL_ZE_IPC_EXCHANGE": "sockets"}, "unset": ["ZE_AFFINITY_MASK"]},
}

# contain the required bs for specific models
specific_models_with_bs: Dict[str, int] = {
    "torch_multimodal_clip": 1,
}

# this list contains the precision we will try to test
#precision_list = ["fp32", "fp16", "amp", "bf16","amp_fp16", "amp_bf16"]
#precision_list = ["amp", "amp_fp16"]
if os.getenv("PRECISIONS"):
    precision_list = os.getenv("PRECISIONS").split("|")
else:
    precision_list = ["fp32", "fp16", "amp", "amp_fp16", "bf16", "amp_bf16"]
print(precision_list)
#precision_list = ["amp"]

# this list contains the backends we will try to test
#backend_list = ["eager", "torchscript"]
backend_list = ["eager"]




# example: {"BERT_pytorch": {"bs": 4, "train": True or None, "inf": True or None}}
models_list: Dict[str, Dict[str, Any]] = {}

pass_model_cnt = 0
fail_model_cnt = 0
not_supported_model_cnt = 0
# example: {"BERT_pytorch": {"result_key": {"pass"  : True}}}
#          {"BERT_pytorch": {"result_key": {"failed": "error message"}}}
#          {"BERT_pytorch": {"result_key": {"Notsupported": "not supported message"}}}
#          {"BERT_pytorch": {"duration": int}}
results_summary: Dict[str, Dict[str, Any]] = {}

def is_eligible_for_test(mode, model):
    if model in models_list_unsupport_eager_mode:
        return False
    return True

def set_the_running_mode(model):
    if mode == 'train':
        models_list[model]['train'] = True
        models_list[model]['inf'] = False
    elif mode == 'inf':
        models_list[model]['train'] = False
        models_list[model]['inf'] = True
    elif mode == 'all':
        models_list[model]['train'] = True
        models_list[model]['inf'] = True
    else:
        raise RuntimeError("[error] no mode is specified, please check env flag TEST_TORCHBENCH_MODE")

def set_the_batch_size(model):
    if model in models_list_unsupport_manual_bs:
        models_list[model]['bs'] = 0
    elif model in specific_models_with_bs:
        models_list[model]['bs'] = specific_models_with_bs[model]
    else:
        models_list[model]['bs'] = run_bs

for model in all_models_list:
    # avoid the recursive deal with model
    if model in models_list:
        continue
    # skip the model which unsupport train/inf in torchbench
    if not is_eligible_for_test(mode, model):
        continue

    models_list[model] = {}

    set_the_running_mode(model)
    set_the_batch_size(model)

# models are eligible to run
num_models = len(models_list.keys())
assert num_models >= 1, "[error] no model is collected"
print('[info] using python executor: ', executor)
print('[info] mode: ', mode)
print('[info] device: ', device)
print('[info] bs: ', run_bs)
print('[info] multi card: ', multi_card)
print('[info] running models list:')
for idx, elem in enumerate(models_list.items()):
    print('[', idx, '] ', elem)

def generate_required_env(model):
    # set the running env
    env = os.environ.copy()

    # remove the env IPEX_SHOW_OPTION for better logging
    env['IPEX_SHOW_OPTION'] = f"0"

    # single card running default
    if not multi_card:
        env['ZE_AFFINITY_MASK'] = f"0"

    if model in models_list_specific_env_flag:
        required_env = models_list_specific_env_flag[model]
        assert "set" in required_env and "unset" in required_env, "check the models_list_specific_env_flag, it is wrongly set"

        # set the required env flag
        for env_name, env_value in required_env["set"].items():
            env[env_name] = env_value

        # remove the env flag
        for env_name in required_env["unset"]:
            del env[env_name]
    return env

def generate_command(model, test_mode='train', precision='fp32', backend='eager'):
    cmd: List[str] = []
    # generate the cmd
    # python -u run.py -d cuda -t train --bs 4 --model "Background_Matting"
    cmd.append(str(executor))
    cmd.append(config_arg)
    cmd.append(bench_file)
    # specify the model
    cmd.append(model)
    cmd.append('-d')
    cmd.append(device)
    cmd.append('-t')
    if test_mode == 'train' and models_list[model]['train']:
        cmd.append('train')
    if test_mode == 'inf' and models_list[model]['inf']:
        cmd.append('eval')
    cmd.append('--bs')
    cmd.append(str(models_list[model]['bs']))
    # disable metrics because it will init pynvml
    cmd.append('--metrics')
    cmd.append('None')
    cmd.append('--precision')
    cmd.append(precision)
    if backend != "eager":
        cmd.append('--backend')
        cmd.append(backend)
    # specify the model
    #cmd.append(model)
    str_cmd = ''
    for elem in cmd:
        str_cmd += elem + ' '
    return cmd, str_cmd.rstrip()

def testing(model, test_mode, precison, backend, count):
    global pass_model_cnt
    global fail_model_cnt
    global not_supported_model_cnt
    result_key = test_mode + precision + backend + '_result'
    if model not in results_summary:
        results_summary[model] = {}
    if result_key not in results_summary[model]:
        results_summary[model][result_key] = {}
    cmd, str_cmd = generate_command(model, test_mode, precision, backend)
    env = generate_required_env(model)
    stderr_msg = None
    stdout_msg = None
    returncode = 0
    is_not_supported = False
    time_start = time.time()
    try:
        # start run the model
        res = subprocess.run(cmd, encoding="utf-8", capture_output=True, check=True, env=env)
        stdout_msg = res.stdout
    except subprocess.CalledProcessError as e:
        returncode = e.returncode
        if returncode != 0:
            stderr_msg = e.stderr
            # check if it is not implemented
            cond1 = unsupported_amp_str in stderr_msg
            cond2 = onlysupportcpu_str in stderr_msg
            cond3 = model in samewithcuda[result_key] if result_key in samewithcuda.keys() else False
            is_not_supported = all([key.lower() in stderr_msg.lower() for key in unsupported_str_list]) or cond1 or cond2 or cond3
    time_end = time.time()
    # set the duration
    duration = round(time_end - time_start, 2)
    results_summary[model]['duration'] = duration

    # for log alignment
    model_space = MODEL_BLANK_SPACE
    model_space = ' ' * (model_space - len(model))
    cmd_space = CMD_BLANK_SPACE
    cmd_space = ' ' * (cmd_space - len(str_cmd))
    if test_mode == 'inf':
        test_mode += ' ' * 2

    # pass
    if returncode == 0:
        print('[', count, '][', test_mode, '][success] pass model: ', model, model_space, 'cmd: ', str_cmd, cmd_space, 'time(s): ', duration)
        pass_model_cnt += 1
        results_summary[model][result_key]['pass'] = True
    else:
        print('[', count, '][', test_mode, '][error]   fail model: ', model, model_space, 'cmd: ', str_cmd, cmd_space, 'time(s): ', duration)
        print(stderr_msg.splitlines()[-1])
        # not supported train or inf
        if is_not_supported:
            #print(f'-----------{model},{result_key}--------')
            not_supported_model_cnt += 1
            results_summary[model][result_key]['Notsupported'] = stderr_msg
            results_summary[model][result_key]['simple_Notsupported'] = stderr_msg.splitlines()[-1]
        else:
            # real failed
            fail_model_cnt += 1
            results_summary[model][result_key]['failed'] = stderr_msg
            results_summary[model][result_key]['simple_failed'] = stderr_msg.splitlines()[-1]

print('[info] running models number: ', num_models)
print('[info] begin test all models......')
global_time_start = time.time()
count = 0
for model in models_list.keys():
    for precision in precision_list:
        for backend in backend_list:
            if models_list[model]['train']:
                testing(model, 'train', precision, backend, count)
                count += 1
            if models_list[model]['inf']:
                testing(model, 'inf', precision, backend, count)
                count += 1

global_time_end = time.time()
global_duration = global_time_end - global_time_start

# summary
print('\n' * 10 + '*' * 50)
print('[info] Detailed Summary:')
for idx, model in enumerate(results_summary.keys()):
    assert model in results_summary, "Fatal error, the model is not in the testing records"

    model_space = MODEL_BLANK_SPACE
    model_space = ' ' * (model_space - len(model))
    for test_mode in ['train', 'inf']:
        for precision in precision_list:
            for backend in backend_list:
                result_key = test_mode + precision + backend + '_result'
                if result_key in results_summary[model]:
                    if 'pass' in results_summary[model][result_key]:
                        print('[', idx, '] model ', model, model_space, test_mode, precision, backend, ' pass')
                    elif 'failed' in results_summary[model][result_key]:
                        print('[', idx, '] model ', model, model_space, test_mode, precision, backend, ' failed:\n', results_summary[model][result_key]['failed'])
                    elif 'Notsupported' in results_summary[model][result_key]:
                        print('[', idx, '] model ', model, model_space, test_mode, precision, backend, ' not supported:\n', results_summary[model][result_key]['Notsupported'])
                    else:
                        raise RuntimeError("model {} is not recorded into the results, check if it is running".format(model))

print('\n' * 10 + '*' * 50)
print('[info] Simplified Summary:')
summary = {}
details = {}
for test_mode in ['train', 'inf']:
    summary[test_mode] = {"error": 0, "passes": 0, "notsupport": 0, "total": 0}
    for precision in precision_list:
        summary[test_mode + "_" + precision] = {"error": 0, "passes": 0, "notsupport": 0, "total": 0}
        for backend in backend_list:
            summary[test_mode + "_" + precision + "_" + backend] = {"error": 0, "passes": 0, "notsupport": 0, "total": 0}
            details[test_mode + "_" + precision + "_" + backend] = {"error": [], "passes": [], "notsupport": []}
for idx, model in enumerate(results_summary.keys()):
    for test_mode in ['train', 'inf']:
        for precision in precision_list:
            for backend in backend_list:
                result_key = test_mode + precision + backend + '_result'
                if result_key in results_summary[model]:
                    if 'pass' in results_summary[model][result_key]:
                        print('[', idx, '] model ', model, test_mode, precision, backend, ' pass')
                        summary[test_mode + "_" + precision + "_" + backend]["passes"] += 1
                        summary[test_mode + "_" + precision + "_" + backend]["total"] += 1
                        summary[test_mode + "_" + precision ]["passes"] += 1
                        summary[test_mode + "_" + precision ]["total"] += 1
                        summary[test_mode ]["passes"] += 1
                        summary[test_mode ]["total"] += 1
                        details[test_mode + "_" + precision + "_" + backend]["passes"].append(model)
                    elif 'failed' in results_summary[model][result_key]:
                        print('[', idx, '] model ', model, test_mode, precision, backend, ' failed')
                        summary[test_mode + "_" + precision + "_" + backend]["error"] += 1
                        summary[test_mode + "_" + precision + "_" + backend]["total"] += 1
                        summary[test_mode + "_" + precision ]["error"] += 1
                        summary[test_mode + "_" + precision ]["total"] += 1
                        summary[test_mode ]["error"] += 1
                        summary[test_mode ]["total"] += 1
                        details[test_mode + "_" + precision + "_" + backend]["error"].append(model)
                    elif 'Notsupported' in results_summary[model][result_key]:
                        print('[', idx, '] model ', model, test_mode, precision, backend, ' not supported')
                        summary[test_mode + "_" + precision + "_" + backend]["notsupport"] += 1
                        summary[test_mode + "_" + precision + "_" + backend]["total"] += 1
                        summary[test_mode + "_" + precision ]["notsupport"] += 1
                        summary[test_mode + "_" + precision ]["total"] += 1
                        summary[test_mode ]["notsupport"] += 1
                        summary[test_mode ]["total"] += 1
                        details[test_mode + "_" + precision + "_" + backend]["notsupport"].append(model)
                    else:
                        raise RuntimeError("model {} is not recorded with the results, check if it run or not".format(model))

# [watch] calculate the pass rate not includes the not supported model
if pass_model_cnt == 0 and fail_model_cnt == 0:
    print('[Error] No pass models or failed models are found')
    sys.exit()

for key in summary.keys():
    if summary[key]["total"] != 0 and summary[key]["total"] - summary[key]["notsupport"] != 0:
        pass_rate = round(summary[key]["passes"] / summary[key]["total"], 3)
        fail_rate = round(summary[key]["error"] / summary[key]["total"], 3)
        pass_rate_wo = round(summary[key]["passes"] / (summary[key]["total"] - summary[key]["notsupport"]), 3)
    elif summary[key]["total"] == 0:
        pass_rate = 0
        fail_rate = 0
        pass_rate_wo = 0
    elif summary[key]["total"] - summary[key]["notsupport"] == 0:
        pass_rate = 0
        fail_rate = 0
        pass_rate_wo = 1
    summary[key]["passrate"] = pass_rate
    summary[key]["failrate"] = fail_rate
    summary[key]["passrate_wo"] = pass_rate_wo
    print(f'[info] {key} pass number = {summary[key]["passes"]}')
    print(f'[info] {key} fail number = {summary[key]["error"]}')
    print(f'[info] {key} not support number = {summary[key]["notsupport"]}')
    print(f'[info] {key} total number = {summary[key]["total"]}')
    print(f'[info] {key} pass rate without not supported= {pass_rate_wo}')
    print(f'[info] {key} pass rate = {pass_rate}')
    print(f'[info] {key} fail rate = {fail_rate}')

print('[info] eligible testing model number = ', num_models)
global_duration = global_duration / 60.0 # unit: min
print('[info] testing total duration = ', round(global_duration, 2), ' min')

with open("summary.log", 'w') as f:
    json.dump(summary, f)
with open("details.log", 'w') as f:
    json.dump(details, f)
with open("results_summary.log", "w") as f:
    json.dump(results_summary, f)

from openpyxl import Workbook
wb = Workbook()
#sheet = wb.active

summary_total = {}
#sheet.append(("modelname", "usecase", "backend", "precision", "result")
ws_s = wb.create_sheet(title="summary")
ws_s.append(("type", "pass", "fail", "notsupport", "total", "passrate w/o unsupported", "passrate", "failrate"))
for key in summary:
    if "eager" not in key and key != "train" and key != "inf":
        continue
    print((key, {summary[key]["passes"]}, {summary[key]["error"]}, {summary[key]["notsupport"]}, {summary[key]["total"]}, summary[key]["passrate_wo"], summary[key]["passrate"], summary[key]["failrate"]))
    ws_s.append((key, summary[key]["passes"], summary[key]["error"], summary[key]["notsupport"], summary[key]["total"], summary[key]["passrate_wo"],  summary[key]["passrate"], summary[key]["failrate"]))
    if key != "train" and key != "inf":
        key1 = key.replace("train_", "").replace("inf_", "")
        if key1 not in summary_total.keys():
            passrate_wo = round(summary[key]["passes"]/(summary[key]["total"] - summary[key]["notsupport"]), 3)
            passrate = round(summary[key]["passes"]/summary[key]["total"], 3)
            failrate = round(summary[key]["error"]/summary[key]["total"], 3)
            summary_total[key1] = {"passes": summary[key]["passes"], "error": summary[key]["error"], "notsupport": summary[key]["notsupport"], "total": summary[key]["total"], "passrate_wo": passrate_wo, "passrate": passrate, "failrate": failrate}
        else:
            summary_total[key1]["passes"] += summary[key]["passes"]
            summary_total[key1]["error"] += summary[key]["error"]
            summary_total[key1]["notsupport"] += summary[key]["notsupport"]
            summary_total[key1]["total"] += summary[key]["total"]
            passrate_wo = round(summary_total[key1]["passes"]/(summary_total[key1]["total"] - summary_total[key1]["notsupport"]), 3)
            passrate = round(summary_total[key1]["passes"]/summary_total[key1]["total"], 3)
            failrate = round(summary_total[key1]["error"]/summary_total[key1]["total"], 3)
            summary_total[key1] = {"passes": summary_total[key1]["passes"], "error": summary_total[key1]["error"], "notsupport": summary_total[key1]["notsupport"], "total": summary_total[key1]["total"], "passrate_wo": passrate_wo, "passrate": passrate, "failrate": failrate}
    elif key == "train" or key == "inf":
        key1 = "total"
        if key1 not in summary_total.keys():
            passrate_wo = round(summary[key]["passes"]/(summary[key]["total"] - summary[key]["notsupport"]), 3)
            passrate = round(summary[key]["passes"]/summary[key]["total"], 3)
            failrate = round(summary[key]["error"]/summary[key]["total"], 3)
            summary_total[key1] = {"passes": summary[key]["passes"], "error": summary[key]["error"], "notsupport": summary[key]["notsupport"], "total": summary[key]["total"], "passrate_wo": passrate_wo, "passrate": passrate, "failrate": failrate}
        else:
            summary_total[key1]["passes"] += summary[key]["passes"]
            summary_total[key1]["error"] += summary[key]["error"]
            summary_total[key1]["notsupport"] += summary[key]["notsupport"]
            summary_total[key1]["total"] += summary[key]["total"]
            passrate_wo = round(summary_total[key1]["passes"]/(summary_total[key1]["total"] - summary_total[key1]["notsupport"]), 3)
            passrate = round(summary_total[key1]["passes"]/summary_total[key1]["total"], 3)
            failrate = round(summary_total[key1]["error"]/summary_total[key1]["total"], 3)
            summary_total[key1] = {"passes": summary_total[key1]["passes"], "error": summary_total[key1]["error"], "notsupport": summary_total[key1]["notsupport"], "total": summary_total[key1]["total"], "passrate_wo": passrate_wo, "passrate": passrate, "failrate": failrate}


for key in summary_total:
    print((key, {summary_total[key]["passes"]}, {summary_total[key]["error"]}, {summary_total[key]["notsupport"]}, {summary_total[key]["total"]}, summary_total[key]["passrate_wo"], summary_total[key]["passrate"], summary_total[key]["failrate"]))
    ws_s.append((key, summary_total[key]["passes"], summary_total[key]["error"], summary_total[key]["notsupport"], summary_total[key]["total"], summary_total[key]["passrate_wo"], summary_total[key]["passrate"], summary_total[key]["failrate"]))


for key in details:
    ws = wb.create_sheet(title=key)
    ws.append(("modelname", "usecase", "backend", "precision", "result", "simple detail", "result details"))
    usecase = key.split("_")[0]
    if "amp_fp16" in key or "amp_bf16" in key:
        precision = key.split("_")[1] + "_" + key.split("_")[2]
        backend = key.split("_")[3]
    else:
        precision = key.split("_")[1]
        backend = key.split("_")[2]
    for result, detail in details[key].items():
        for mymodel in detail:
            result_key = usecase + precision + backend + "_result"
            if result == "error":
                result_detail = results_summary[mymodel][result_key]['failed']
                simple_detail = results_summary[mymodel][result_key]['simple_failed']
            elif result == "notsupport":
                result_detail = results_summary[mymodel][result_key]['Notsupported']
                simple_detail = results_summary[mymodel][result_key]['simple_Notsupported']
            else:
                result_detail = ""
                simple_detail = ""
            ws.append((mymodel, usecase, backend, precision, result, simple_detail, result_detail))

wb.save("torchbench.xlsx")



