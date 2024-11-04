# runn all models usage:
# under the torchbench folder and python -u test_torchbench.py
# run single model usage:
# bs is 0 means use the default value
# python -u run.py -d xpu -t train --bs 0 --metrics None Model_name
# bs is 4 means set the bs to 4
# python -u run.py -d xpu -t train --bs 4 --metrics None Model_name

import os
import subprocess
from typing import List, Dict, Any
import sys
import time
import json


# config
run_bs = 4 # default 4, it is used the batch size which model can be set
executor = sys.executable
device = 'cuda' # default xpu
unsupported_str_list = ["train", "notimplement", "support"]
unsupported_amp_str = "error: unrecognized arguments: --amp"
onlysupportcpu_str = "The eval test only supports CPU"
mode = os.getenv('TEST_HF_MODE')
os.environ["TRUST_REMOTE_CODE"] = "True"
# default is testing train/inf
if mode is None:
    mode = 'all'
mode = 'train'
assert type(mode) is str and mode.lower() in ['train', 'inf', 'all'], "please specify the TEST_TORCHVISION_MODE to be train/inf/all"
bench_file = 'train.py'
config_arg = '-u' # python -u xxxx
multi_card = False # default single card running

# log print
MODEL_BLANK_SPACE = 50
CMD_BLANK_SPACE = 150

# all models scope in torchvision
all_models_dict = {
    "image_caption": {
        "train": [
            "microsoft/git-base",
            "nlpconnect/vit-gpt2-image-captioning",
            "Salesforce/blip-image-captioning-large",
            "Salesforce/blip-image-captioning-base",
            "Salesforce/blip2-opt-2.7b",
            "facebook/nougat-base",
            "google/pix2struct-large",
            "Salesforce/blip2-flan-t5-xl",
            "Salesforce/blip2-opt-2.7b-coco",
            "google/pix2struct-textcaps-base",
            "microsoft/kosmos-2-patch14-224",
            "alibaba-damo/mgp-str-base",
            "unum-cloud/uform-gen2-qwen-500m"
            ],
        "inf": [
            ]
        },
    "image_classification": {
        "train": [
            "google/vit-base-patch16-224-in21k",
            "microsoft/resnet-50",
            "google/vit-base-patch16-224",
            "microsoft/beit-base-patch16-224-pt22k-ft22k",
            "Falconsai/nsfw_image_detection",
            "AdamCodd/vit-base-nsfw-detector",
            "microsoft/swinv2-tiny-patch4-window16-256",
            "google/mobilenet_v1_0.75_192",
            "google/vit-large-patch32-384",
            "facebook/convnextv2-tiny-1k-224",
            "microsoft/dit-base-finetuned-rvlcdip",
            "microsoft/resnet-18",
            "google/mobilenet_v2_1.0_224",
            "nvidia/mit-b0",
            "google/vit-large-patch16-224",
            "microsoft/swin-base-patch4-window12-384",
            "facebook/convnext-base-224-22k",
            "nvidia/mit-b5",
            "microsoft/swin-tiny-patch4-window7-224",
            "nvidia/mit-b1",
            ],
        "inf": [
            "microsoft/resnet-50",
            "google/vit-base-patch16-224",
            "Falconsai/nsfw_image_detection",
            "nateraw/vit-age-classifier",
            "microsoft/beit-base-patch16-224-pt22k-ft22k",
            "microsoft/resnet-18",
            "nvidia/mit-b0",
            "google/vit-base-patch16-384",
            "google/vit-large-patch16-224",
            "facebook/deit-base-distilled-patch16-224",
            "google/vit-base-patch32-384",
            "microsoft/resnet-101",
            "microsoft/cvt-13",
            "google/vit-large-patch16-384",
            "microsoft/beit-base-patch16-224",
            "microsoft/beit-large-patch16-512",
            "microsoft/resnet-152",
            "google/mobilenet_v2_1.0_224",
            "google/efficientnet-b7",
            "facebook/convnext-base-224",
            ]
        },
    "question_answering": {
        "train": [
            "distilbert/distilbert-base-uncased",
            "deepset/roberta-base-squad2",
            "deepset/bert-large-uncased-whole-word-masking-squad2",
            "distilbert/distilbert-base-cased-distilled-squad",
            "deepset/roberta-large-squad2",
            "sjrhuschlee/flan-t5-base-squad2",
            "deepset/minilm-uncased-squad2",
            "deepset/tinyroberta-squad2",
            "distilbert/distilbert-base-uncased-distilled-squad",
            "uer/roberta-base-chinese-extractive-qa",
            "deepset/deberta-v3-large-squad2",
            "deepset/deberta-v3-base-squad2",
            "deepset/xlm-roberta-large-squad2",
            "pedramyazdipoor/persian_xlm_roberta_large",
            "DeepMount00/Gemma_QA_ITA_v3",
            "VietAI/vit5-base",
            "deepset/xlm-roberta-base-squad2-distilled",
            "pierreguillou/bert-base-cased-squad-v1.1-portuguese",
            "deepset/xlm-roberta-base-squad2",
            "VietAI/vit5-large",
            ],
        "inf": [
            ]
        },
    "summarization": {
        "train": [
            "google-t5/t5-small",
            "facebook/bart-large-cnn",
            "sshleifer/distilbart-cnn-12-6",
            "philschmid/bart-large-cnn-samsum",
            "google/pegasus-xsum",
            "Falconsai/medical_summarization",
            "transformer3/H2-keywordextractor",
            "sshleifer/distilbart-cnn-12-3",
            "google/pegasus-multi_news",
            "nandakishormpai/t5-small-machine-articles-tag-generation",
            "slauw87/bart_summarisation",
            "cnicu/t5-small-booksum",
            "knkarthick/MEETING_SUMMARY",
            "jotamunz/billsum_tiny_summarization",
            "sshleifer/distilbart-cnn-6-6",
            "human-centered-summarization/financial-summarization-pegasus",
            "google/pegasus-cnn_dailymail",
            "IlyaGusev/mbart_ru_sum_gazeta",
            "lidiya/bart-large-xsum-samsum",
            "gsarti/it5-base-news-summarization",
            ],
        "inf": [
            "facebook/bart-large-cnn",
            "sshleifer/distilbart-cnn-12-6",
            "philschmid/bart-large-cnn-samsum",
            "google/pegasus-xsum",
            "Falconsai/text_summarization",
            "suriya7/bart-finetuned-text-summarization",
            "human-centered-summarization/financial-summarization-pegasus",
            "cointegrated/rut5-base-absum",
            "Einmalumdiewelt/T5-Base_GNAD",
            "google/bigbird-pegasus-large-arxiv",
            "sshleifer/distilbart-cnn-6-6",
            "google/pegasus-cnn_dailymail",
            "jotamunz/billsum_tiny_summarization",
            "facebook/bart-large-xsum",
            ]
        },
    "text_classification": {
        "train": [
            "distilbert/distilbert-base-uncased_version_1",
            "distilbert/distilbert-base-uncased-finetuned-sst-2-english",
            "cardiffnlp/twitter-roberta-base-sentiment-latest",
            "avichr/heBERT_sentiment_analysis",
            "cross-encoder/ms-marco-TinyBERT-L-2-v2",
            "papluca/xlm-roberta-base-language-detection",
            "cardiffnlp/twitter-roberta-base-sentiment",
            "Ashishkr/query_wellformedness_score",
            "cross-encoder/ms-marco-MiniLM-L-6-v2",
            "cross-encoder/ms-marco-MiniLM-L-12-v2",
            "martin-ha/toxic-comment-model",
            "BAAI/bge-reranker-v2-m3",
            ],
        "inf": [
            "ProsusAI/finbert",
            "distilbert/distilbert-base-uncased-finetuned-sst-2-english",
            "cardiffnlp/twitter-roberta-base-sentiment-latest",
            "SamLowe/roberta-base-go_emotions",
            "j-hartmann/emotion-english-distilroberta-base",
            "mrm8488/distilroberta-finetuned-financial-news-sentiment-analysis",
            "nlptown/bert-base-multilingual-uncased-sentiment",
            "cardiffnlp/twitter-roberta-base-sentiment",
            "papluca/xlm-roberta-base-language-detection",
            "lxyuan/distilbert-base-multilingual-cased-sentiments-student",
            "OpenAssistant/reward-model-deberta-v3-large-v2",
            "vectara/hallucination_evaluation_model",
            "cardiffnlp/twitter-xlm-roberta-base-sentiment",
            "yiyanghkust/finbert-tone",
            "FacebookAI/roberta-large-mnli",
            "facebook/fasttext-language-identification",            #Fail due to repo issue
            "unitary/toxic-bert",
            "finiteautomata/bertweet-base-sentiment-analysis",
            "BAAI/bge-reranker-base",
            "notdiamond/notdiamond-0001",
            ]
        },
    "token_classification": {
        "train": [
            "distilbert/distilbert-base-uncased_version_2",
            "dslim/bert-base-NER",
            "FacebookAI/xlm-roberta-large-finetuned-conll03-english",
            "blaze999/Medical-NER",
            "ml6team/keyphrase-extraction-distilbert-inspec",
            "KoichiYasuoka/bert-base-thai-upos",
            "akdeniz27/bert-base-turkish-cased-ner",
            "51la5/roberta-large-NER",
            "Babelscape/wikineural-multilingual-ner",
            "dslim/bert-large-NER",
            "Davlan/bert-base-multilingual-cased-ner-hrl",
            "pierreguillou/ner-bert-base-cased-pt-lenerbr",
            "KB/bert-base-swedish-cased-ner",
            "NlpHUST/ner-vietnamese-electra-base",
            "Jean-Baptiste/camembert-ner",
            "SenswiseData/bert_cased_ner",
            "dslim/distilbert-NER",
            "ml6team/keyphrase-extraction-kbir-inspec",
            "ugaray96/biobert_ncbi_disease_ner",
            "Jean-Baptiste/roberta-large-ner-english",
            "NbAiLab/nb-bert-base-ner",
            ],
        "inf": [
            "dslim/bert-base-NER",
            "FacebookAI/xlm-roberta-large-finetuned-conll03-english",
            "blaze999/Medical-NER",
            "ckiplab/bert-base-chinese-ner",
            "Babelscape/wikineural-multilingual-ner",
            "dslim/bert-large-NER",
            "xlm-roberta-large-finetuned-conll03-english",
            "KoichiYasuoka/bert-base-thai-upos",
            "pierreguillou/ner-bert-base-cased-pt-lenerbr",
            "Davlan/bert-base-multilingual-cased-ner-hrl",
            "KB/bert-base-swedish-cased",
            "ckiplab/albert-tiny-chinese-ws",
            ]
        },
    "translation": {
        "train": [
            "google-t5/t5-small_version_1",
            "google-t5/t5-base",
            "google-t5/t5-large",
            "Helsinki-NLP/opus-mt-mul-en",
            "facebook/nllb-200-distilled-600M",
            "Helsinki-NLP/opus-mt-de-en",
            "google-t5/t5-3b",
            "optimum/t5-small",
            "facebook/wmt19-ru-en",
            "Helsinki-NLP/opus-mt-fr-de",
            ],
        "inf": [
            ]
        },
    "depth_estimation": {
        "train": [
            ],
        "inf": [
            "ProsusAI/finbert_version_1",
            "LiheYoung/depth-anything-small-hf",
            "vinvino02/glpn-kitti",
            "facebook/dpt-dinov2-base-kitti",
            "depth-anything/Depth-Anything-V2-Small-hf",
            "LiheYoung/depth-anything-large-hf",
            "LiheYoung/depth-anything-base-hf",
            "vinvino02/glpn-nyu",
            "depth-anything/Depth-Anything-V2-Base-hf",
            "facebook/dpt-dinov2-large-nyu",
            "facebook/dpt-dinov2-small-kitti",
            ]
        },
    "image_segmentation": {
        "train": [
            ],
        "inf": [
            "mattmdjaga/segformer_b2_clothes",
            "nvidia/segformer-b0-finetuned-ade-512-512",
            "jonathandinu/face-parsing",
            "sayeed99/segformer_b3_clothes",
            "nvidia/segformer-b2-finetuned-ade-512-512",
            "nvidia/segformer-b3-finetuned-cityscapes-1024-1024",
            "nvidia/segformer-b4-finetuned-ade-512-512",
            "nvidia/segformer-b3-finetuned-ade-512-512",
            "nvidia/segformer-b1-finetuned-cityscapes-1024-1024",
            "sayeed99/segformer-b3-fashion",
            "nvidia/segformer-b5-finetuned-cityscapes-1024-1024",
            "microsoft/beit-large-finetuned-ade-640-640",
            ]
        },
    "image_to_text": {
        "train": [
            ],
        "inf": [
            "microsoft/trocr-base-handwritten",       # Failed for tools reason
            "Salesforce/blip-image-captioning-large_version_1",
            "Salesforce/blip-image-captioning-base_version_1",
            "Salesforce/blip2-opt-2.7b_version_1",                # Failed for tools reason but have fix didn't land
            "microsoft/trocr-small-handwritten",
            "facebook/nougat-base_version_1",
            "Salesforce/instructblip-vicuna-7b",
            "microsoft/trocr-large-printed",
            "Salesforce/blip2-flan-t5-xl_version_1",
            "microsoft/trocr-base-printed",
            "alibaba-damo/mgp-str-base_version_1",
            "nlpconnect/vit-gpt2-image-captioning_version_1",
            ]
        },
    "object_detection": {
        "train": [
            ],
        "inf": [
            "facebook/detr-resnet-101",   #Failed to skip
            "facebook/detr-resnet-50",    #Failed to skip
            "hustvl/yolos-tiny",
            "SenseTime/deformable-detr",
            "Aryn/deformable-detr-DocLayNet",
            "nickmuchi/yolos-small-finetuned-license-plate-detection",
            "SenseTime/deformable-detr-with-box-refine",
            "hustvl/yolos-base",
            "microsoft/conditional-detr-resnet-50",
            "PekingU/rtdetr_r50vd",
            "facebook/detr-resnet-101-dc5",
            "PekingU/rtdetr_r101vd_coco_o365",
            ]
        },
    "sentence_similarity": {
        "train": [
            ],
        "inf": [
            "sentence-transformers/all-MiniLM-L6-v2",
            "sentence-transformers/all-MiniLM-L12-v2",
            "sentence-transformers/all-mpnet-base-v2",
            "sentence-transformers/paraphrase-MiniLM-L6-v2",
            "sentence-transformers/bert-base-nli-mean-tokens",
            "sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2",
            "sentence-transformers/multi-qa-MiniLM-L6-cos-v1",
            "sentence-transformers/msmarco-bert-base-dot-v5",
            "Alibaba-NLP/gte-large-en-v1.5",
            "sentence-transformers/distiluse-base-multilingual-cased-v2",
            "sentence-transformers/multi-qa-mpnet-base-dot-v1",
            "sentence-transformers/paraphrase-multilingual-mpnet-base-v2",
            ]
        },
    "text_zero_shot_classification": {
        "train": [
            ],
        "inf": [
            "tasksource/deberta-small-long-nli",
            "facebook/bart-large-mnli",
            "MoritzLaurer/DeBERTa-v3-base-mnli-fever-anli",
            "cross-encoder/nli-roberta-base",
            "MoritzLaurer/deberta-v3-large-zeroshot-v2.0",
            "MoritzLaurer/mDeBERTa-v3-base-xnli-multilingual-nli-2mil7",
            "cross-encoder/nli-deberta-v3-base",
            "cross-encoder/nli-deberta-base",
            "joeddav/xlm-roberta-large-xnli",
            "cross-encoder/nli-distilroberta-base",
            "cross-encoder/nli-MiniLM2-L6-H768",
            "sileod/deberta-v3-base-tasksource-nli",
            ]
        }
    }
for key in all_models_dict.keys():
    print(key)
common_args_dict = {
                "classification": {
                    "lr": "0.1",
                    "datapath": "imagenet"
                    },
                "detection": {
                    "lr": "0.02",
                    "datapath" : "coco"
                    },
                "segmentation": {
                    "lr": "0.02",
                    "datapath" : "coco",
                    },
                "video": {
                    "datapath": "tiny-Kinetics-400",
                    "lr": "0.02",
                    },
                "optical_flow": {
                    "lr": "0.0004",
                    },
                "quantization": {
                    "lr": "0.1",
                    "datapath": "imagenet",
                    },
                }
extra_category_args = {
                "classification": {
                    },
                "detection": {
                    },
                "segmentation": {
                    },
                "video": {
                    "inf": ["--kinetics-version", "400", "--batch-size", "2", "--clip-len", "1", "--frame-rate", "1"],
                    "train": ["--kinetics-version", "400", "--batch-size", "2", "--clip-len", "1", "--frame-rate", "1"],
                    #"inf": "--kinetics-version 400 --batch-size 2 --clip-len 1 --frame-rate 1",
                    #"train": "--kinetics-version 400 --batch-size 2 --clip-len 1 --frame-rate 1",
                    },
                "optical_flow": {
                    "inf": ["--val-dataset", "kitti", "--batch-size", "1", "--dataset-root", "."],
                    "train": ["--train-dataset", "kitti", "--val-dataset", "kitti", "--batch-size", "2", "--dataset-root", "."],
                    },
                "quantization": {
                    "inf": ["--qbackend", "fbgemm", "--batch-size", "1",],
                    "train": ["--post-training-quantize", "--qbackend", "fbgemm", "--batch-size", "1"],
                    },
                }
extra_args = {
        "ssd300_vgg16": {
            "train": ["--batch-size", "4"],
            }
        }

#all_models_list = [
#            "BERT_pytorch"]

# This list contains the model which is unsupported to set the maunal bs
# so set the bs 0 in command
models_list_unsupport_manual_bs = [
    "basic_gnn_edgecnn",
    "basic_gnn_gcn",
    "basic_gnn_gin",
    "basic_gnn_sage",
    "Background_Matting",
    "functorch_maml_omniglot",
    "maml",
    "maml_omniglot",
    "pytorch_CycleGAN_and_pix2pix",
    "pytorch_stargan",
    "soft_actor_critic",
    "speech_transformer",
    "stable_diffusion_text_encoder",
    "stable_diffusion_unet",
    "torch_multimodal_clip",
    "vision_maskrcnn",
    "pyhpc_turbulent_kinetic_energy",
    "drq",
]

# This list contains the model which is unsupported to train
# the train function is not implemented by torchbench
models_list_unsupport_eager_mode = [
    "simple_gpt",
    "simple_gpt_tp_manual",
]

# This list contains the model which needs to set/unset the specific env flag
# "model": {"set": {"env_name":"env_value"}, "unset":["env_name"]}
models_list_specific_env_flag = {
    "moco":            {"set": {"CCL_ZE_IPC_EXCHANGE": "sockets"}, "unset": ["ZE_AFFINITY_MASK"]},
    "vision_maskrcnn": {"set": {"CCL_ZE_IPC_EXCHANGE": "sockets"}, "unset": ["ZE_AFFINITY_MASK"]},
}

# contain the required bs for specific models
specific_models_with_bs: Dict[str, int] = {
    "torch_multimodal_clip": 1,
}

# this list contains the precision we will try to test
#precision_list = ["fp32", "fp16", "amp", "bf16","amp_fp16", "amp_bf16"]
#precision_list = ["fp32", "fp16", "bf16", "amp"]
if os.getenv("PRECISIONS"):
    precision_list = os.getenv("PRECISIONS").split("|")
else:
    precision_list = ["fp32", "fp16", "bf16"]
print(precision_list)
#precision_list = ["amp"]

# this list contains the backends we will try to test
#backend_list = ["eager", "torchscript"]
#backend_list = ["eager", "torch_compile"]
if os.getenv("BACKENDS"):
    backend_list1 = os.getenv("BACKENDS").split("|")
    backend_list = []
    if "torchscript" in backend_list1:
        for backend in backend_list1:
            backend1 = backend.replace("torchscript", "torch_compile")
            backend_list.append(backend1)
    else:
        backend_list = backend_list1        
else:
    backend_list = ["eager"]
print(backend_list)





# example: {"BERT_pytorch": {"bs": 4, "train": True or None, "inf": True or None}}
models_list: Dict[str, Dict[str, Any]] = {}

pass_model_cnt = 0
fail_model_cnt = 0
not_supported_model_cnt = 0
# example: {"BERT_pytorch": {"result_key": {"pass"  : True}}}
#          {"BERT_pytorch": {"result_key": {"failed": "error message"}}}
#          {"BERT_pytorch": {"result_key": {"Notsupported": "not supported message"}}}
#          {"BERT_pytorch": {"duration": int}}
results_summary: Dict[str, Dict[str, Any]] = {}

def is_eligible_for_test(mode, model):
    if model in models_list_unsupport_eager_mode:
        return False
    return True

#def set_the_running_mode(model):
#    if mode == 'train':
#        models_list[model]['train'] = True
#        models_list[model]['inf'] = False
#    elif mode == 'inf':
#        models_list[model]['train'] = False
#        models_list[model]['inf'] = True
#    elif mode == 'all':
#        models_list[model]['train'] = True
#        models_list[model]['inf'] = True
#    else:
#        raise RuntimeError("[error] no mode is specified, please check env flag TEST_TORCHBENCH_MODE")

def set_the_batch_size(model):
    if model in models_list_unsupport_manual_bs:
        models_list[model]['bs'] = 0
    elif model in specific_models_with_bs:
        models_list[model]['bs'] = specific_models_with_bs[model]
    else:
        models_list[model]['bs'] = run_bs

for category, mode_list in all_models_dict.items():
    # avoid the recursive deal with model
    for mode, model_list in mode_list.items():
        for model in model_list:
            if not is_eligible_for_test(mode, model):
                continue
            if model in models_list:
                models_list[model][mode] = True
                print(f'already has {models_list[model]["category"]} but add {category}')
            else:
                models_list[model] = {"category": category, 'train': False, 'inf': False}
                models_list[model][mode] = True

            #set_the_running_mode(model)
            set_the_batch_size(model)

# models are eligible to run
num_models = len(models_list.keys())
assert num_models >= 1, "[error] no model is collected"
print('[info] using python executor: ', executor)
print('[info] mode: ', mode)
print('[info] device: ', device)
print('[info] bs: ', run_bs)
print('[info] multi card: ', multi_card)
print('[info] running models list:')
for idx, elem in enumerate(models_list.items()):
    print('[', idx, '] ', elem)

def generate_required_env(model):
    # set the running env
    env = os.environ.copy()

    # remove the env IPEX_SHOW_OPTION for better logging
    env['IPEX_SHOW_OPTION'] = f"0"

    # single card running default
    if not multi_card:
        env['ZE_AFFINITY_MASK'] = f"0"

    if model in models_list_specific_env_flag:
        required_env = models_list_specific_env_flag[model]
        assert "set" in required_env and "unset" in required_env, "check the models_list_specific_env_flag, it is wrongly set"

        # set the required env flag
        for env_name, env_value in required_env["set"].items():
            env[env_name] = env_value

        # remove the env flag
        for env_name in required_env["unset"]:
            del env[env_name]
    return env

def generate_command(model, category, test_mode='train', precision='fp32', backend='eager'):
    cmd: List[str] = []
    global bench_file
    # generate the cmd
    # python -u references/detection/train.py --lr 0.02 --epochs  1  --device cpu --data-path /home/gta/wliao2/torchvision/coco --model fasterrcnn_mobilenet_v3_large_320_fpn --aspect-ratio-group-factor 3 --weights-backbone MobileNet_V3_Large_Weights.IMAGENET1K_V1
    cmd.append(str(executor))
    cmd.append(config_arg)
    if test_mode == 'train':
        bench_path = "training/"
    elif test_mode == 'inf':
        bench_path = "inference/"
    bench_file = "test_" + category + ".py"
    cmd.append(bench_path + bench_file)
    print(f'cmd is {cmd}')
    # specify the model
    cmd.append('--model')
    cmd.append(model.replace("_version_1", "").replace("_version_2", ""))
    # disable metrics because it will init pynvml
    #if precision == "amp":
    #    cmd.append('--amp')
    #if precision != "default":
    cmd.append('--precision')
    cmd.append(precision)
    #if backend != "eager":
    cmd.append('--backend')
    cmd.append(backend)
    #if extra_category_args[category]:
    #    if test_mode in extra_category_args[category]:
    #        cmd.extend(extra_category_args[category][test_mode])
    #if model in extra_args.keys() and test_mode in extra_args[model].keys():
    #    cmd.extend(extra_args[model][test_mode])
    # specify the model
    #cmd.append(model)
    str_cmd = ''
    print(cmd)
    for elem in cmd:
        str_cmd += elem + ' '
    return cmd, str_cmd.rstrip()

def testing(model, category, test_mode, precison, backend, count):
    global pass_model_cnt
    global fail_model_cnt
    global not_supported_model_cnt
    result_key = test_mode + precision + backend + '_result'
    if model not in results_summary:
        results_summary[model] = {}
    if result_key not in results_summary[model]:
        results_summary[model][result_key] = {}
    print(f'{category},{model}')
    cmd, str_cmd = generate_command(model, category, test_mode, precision, backend)
    env = generate_required_env(model)
    stderr_msg = None
    returncode = 0
    is_not_supported = False
    time_start = time.time()
    try:
        # start run the model
        print(f'begin to run cmd" {cmd} {str_cmd}')
        subprocess.run(cmd, encoding="utf-8", capture_output=True, check=True, env=env)
        print("finished")
    except subprocess.CalledProcessError as e:
        returncode = e.returncode
        if returncode != 0:
            stderr_msg = e.stderr
            # check if it is not implemented
            cond1 = unsupported_amp_str in stderr_msg
            cond2 = onlysupportcpu_str in stderr_msg
            #cond3 = model in samewithcuda[result_key] if result_key in samewithcuda.keys() else False
            cond3 = False
            is_not_supported = all([key.lower() in stderr_msg.lower() for key in unsupported_str_list]) or cond1 or cond2 or cond3
    time_end = time.time()
    # set the duration
    duration = round(time_end - time_start, 2)
    results_summary[model]['duration'] = duration

    # for log alignment
    model_space = MODEL_BLANK_SPACE
    model_space = ' ' * (model_space - len(model))
    cmd_space = CMD_BLANK_SPACE
    cmd_space = ' ' * (cmd_space - len(str_cmd))
    if test_mode == 'inf':
        test_mode += ' ' * 2
    results_summary[model][result_key]['category'] = category

    # pass
    if returncode == 0:
        print('[', count, '][', test_mode, '][success] pass model: ', model, model_space, 'cmd: ', str_cmd, cmd_space, 'time(s): ', duration)
        pass_model_cnt += 1
        results_summary[model][result_key]['pass'] = True
    else:
        print('[', count, '][', test_mode, '][error]   fail model: ', model, model_space, 'cmd: ', str_cmd, cmd_space, 'time(s): ', duration)
        simple_error = "no error or fail found in stderr"
        for i in range(1, len(stderr_msg.splitlines())):
            #print(f'i is {i} and msg={stderr_msg.splitlines()[-i]}')
            if "error" in stderr_msg.splitlines()[-i].lower() or "fail" in stderr_msg.splitlines()[-i].lower() or "fault" in stderr_msg.splitlines()[-i].lower():
                simple_error = stderr_msg.splitlines()[-i]
                print(simple_error)
                break
        # not supported train or inf
        if is_not_supported:
            #print(f'-----------{model},{result_key}--------')
            not_supported_model_cnt += 1
            results_summary[model][result_key]['Notsupported'] = stderr_msg
            results_summary[model][result_key]['simple_Notsupported'] = simple_error
        else:
            # real failed
            fail_model_cnt += 1
            results_summary[model][result_key]['failed'] = stderr_msg
            results_summary[model][result_key]['simple_failed'] = simple_error

print('[info] running models number: ', num_models)
print('[info] begin test all models......')
global_time_start = time.time()
count = 0
for model in models_list.keys():
    for precision in precision_list:
        for backend in backend_list:
            if 'train' in models_list[model].keys() and models_list[model]['train']:
                testing(model, models_list[model]['category'], 'train', precision, backend, count)
                count += 1
            if 'inf' in models_list[model].keys() and models_list[model]['inf']:
                testing(model, models_list[model]['category'], 'inf', precision, backend, count)
                count += 1

global_time_end = time.time()
global_duration = global_time_end - global_time_start

# summary
print('\n' * 10 + '*' * 50)
print('[info] Detailed Summary:')
for idx, model in enumerate(results_summary.keys()):
    assert model in results_summary, "Fatal error, the model is not in the testing records"

    model_space = MODEL_BLANK_SPACE
    model_space = ' ' * (model_space - len(model))
    for test_mode in ['train', 'inf']:
        if not all_models_dict[category][test_mode]:
            continue
        for precision in precision_list:
            for backend in backend_list:
                result_key = test_mode + precision + backend + '_result'
                if result_key in results_summary[model]:
                    if 'pass' in results_summary[model][result_key]:
                        print('[', idx, '] model ', model, model_space, test_mode, precision, backend, ' pass')
                    elif 'failed' in results_summary[model][result_key]:
                        print('[', idx, '] model ', model, model_space, test_mode, precision, backend, ' failed:\n', results_summary[model][result_key]['failed'])
                    elif 'Notsupported' in results_summary[model][result_key]:
                        print('[', idx, '] model ', model, model_space, test_mode, precision, backend, ' not supported:\n', results_summary[model][result_key]['Notsupported'])
                    else:
                        raise RuntimeError("model {} is not recorded into the results, check if it is running".format(model))

print('\n' * 10 + '*' * 50)
print('[info] Simplified Summary:')
with open("results_summary.log", "w") as f:
    json.dump(results_summary, f)

summary = {}
details = {}
#import pdb;pdb.set_trace()
for category in all_models_dict.keys():
    summary[category] = {"error": 0, "passes": 0, "notsupport": 0, "total": 0}
    for test_mode in ['train', 'inf']:
        if not all_models_dict[category][test_mode]:
            continue
        summary[category + "_" + test_mode] = {"error": 0, "passes": 0, "notsupport": 0, "total": 0}
        for precision in precision_list:
            summary[category + "_" + test_mode + "_" + precision] = {"error": 0, "passes": 0, "notsupport": 0, "total": 0}
            for backend in backend_list:
                summary[category + "_" + test_mode + "_" + precision + "_" + backend] = {"error": 0, "passes": 0, "notsupport": 0, "total": 0}
                details[category + "_" + test_mode + "_" + precision + "_" + backend] = {"error": [], "passes": [], "notsupport": []}
for idx, model in enumerate(results_summary.keys()):
    for test_mode in ['train', 'inf']:
        for precision in precision_list:
            for backend in backend_list:
                result_key = test_mode + precision + backend + '_result'
                category = models_list[model]['category']
            if result_key in results_summary[model]:
                if 'pass' in results_summary[model][result_key]:
                    print('[', idx, '] model ', category, model, test_mode, precision, backend, ' pass')
                    summary[category + "_" + test_mode + "_" + precision + "_" + backend]["passes"] += 1
                    summary[category + "_" + test_mode + "_" + precision + "_" + backend]["total"] += 1
                    summary[category + "_" + test_mode + "_" + precision ]["passes"] += 1
                    summary[category + "_" + test_mode + "_" + precision ]["total"] += 1
                    summary[category + "_" + test_mode ]["passes"] += 1
                    summary[category + "_" + test_mode ]["total"] += 1
                    summary[category]["passes"] += 1
                    summary[category]["total"] += 1
                    details[category + "_" + test_mode + "_" + precision + "_" + backend]["passes"].append(model)
                elif 'failed' in results_summary[model][result_key]:
                    print('[', idx, '] model ', category, model, test_mode, precision, backend, ' fail')
                    summary[category + "_" + test_mode + "_" + precision + "_" + backend]["error"] += 1
                    summary[category + "_" + test_mode + "_" + precision + "_" + backend]["total"] += 1
                    summary[category + "_" + test_mode + "_" + precision ]["error"] += 1
                    summary[category + "_" + test_mode + "_" + precision ]["total"] += 1
                    summary[category + "_" + test_mode ]["error"] += 1
                    summary[category + "_" + test_mode ]["total"] += 1
                    summary[category]["error"] += 1
                    summary[category]["total"] += 1
                    details[category + "_" + test_mode + "_" + precision + "_" + backend]["error"].append(model)
                elif 'Notsupported' in results_summary[model][result_key]:
                    print('[', idx, '] model ', category, model, test_mode, precision, backend, ' not supported')
                    summary[category + "_" + test_mode + "_" + precision + "_" + backend]["notsupport"] += 1
                    summary[category + "_" + test_mode + "_" + precision + "_" + backend]["total"] += 1
                    summary[category + "_" + test_mode + "_" + precision ]["notsupport"] += 1
                    summary[category + "_" + test_mode + "_" + precision ]["total"] += 1
                    summary[category + "_" + test_mode ]["notsupport"] += 1
                    summary[category + "_" + test_mode ]["total"] += 1
                    summary[category]["notsupport"] += 1
                    summary[category]["total"] += 1
                    details[category + "_" + test_mode + "_" + precision + "_" + backend]["notsupport"].append(model)
                else:
                    raise RuntimeError("model {} is not recorded with the results, check if it run or not".format(model))

# [watch] calculate the pass rate not includes the not supported model
if pass_model_cnt == 0 and fail_model_cnt == 0:
    print('[Error] No pass models or failed models are found')
    sys.exit()

for key in summary.keys():
    if summary[key]["total"] != 0 and summary[key]["total"] - summary[key]["notsupport"] != 0:
        pass_rate = round(summary[key]["passes"] / summary[key]["total"], 3)
        fail_rate = round(summary[key]["error"] / summary[key]["total"], 3)
        pass_rate_wo = round(summary[key]["passes"] / (summary[key]["total"] - summary[key]["notsupport"]), 3)
    elif summary[key]["total"] == 0:
        pass_rate = 0
        fail_rate = 0
        pass_rate_wo = 0
    elif summary[key]["total"] - summary[key]["notsupport"] == 0:
        pass_rate = 0
        fail_rate = 0
        pass_rate_wo = 1
    summary[key]["passrate"] = pass_rate
    summary[key]["failrate"] = fail_rate
    summary[key]["passrate_wo"] = pass_rate_wo
    print(f'[info] {key} pass number = {summary[key]["passes"]}')
    print(f'[info] {key} fail number = {summary[key]["error"]}')
    print(f'[info] {key} not support number = {summary[key]["notsupport"]}')
    print(f'[info] {key} total number = {summary[key]["total"]}')
    print(f'[info] {key} pass rate without not supported= {pass_rate_wo}')
    print(f'[info] {key} pass rate = {pass_rate}')
    print(f'[info] {key} fail rate = {fail_rate}')

print('[info] eligible testing model number = ', num_models)
global_duration = global_duration / 60.0 # unit: min
print('[info] testing total duration = ', round(global_duration, 2), ' min')

with open("summary.log", 'w') as f:
    json.dump(summary, f)
with open("details.log", 'w') as f:
    json.dump(details, f)

from openpyxl import Workbook, load_workbook
wb = Workbook()
#sheet = wb.active

summary_total = {}
#sheet.append(("modelname", "usecase", "backend", "precision", "result")
ws_s = wb.create_sheet(title="summary")
ws_s.append(("type", "pass", "fail", "notsupport", "total", "passrate w/o unsupported", "passrate", "failrate"))
for key in summary:
    if "eager" not in key and key != "train" and key != "inf":
        continue
    print((key, {summary[key]["passes"]}, {summary[key]["error"]}, {summary[key]["notsupport"]}, {summary[key]["total"]}, summary[key]["passrate_wo"], summary[key]["passrate"], summary[key]["failrate"]))
    ws_s.append((key, summary[key]["passes"], summary[key]["error"], summary[key]["notsupport"], summary[key]["total"], summary[key]["passrate_wo"],  summary[key]["passrate"], summary[key]["failrate"]))
    if key != "train" and key != "inf":
        key1 = key.replace("train_", "").replace("inf_", "")
        if key1 not in summary_total.keys():
            if summary[key]["total"] - summary[key]["notsupport"] != 0:
                passrate_wo = round(summary[key]["passes"]/(summary[key]["total"] - summary[key]["notsupport"]), 3)
            else:
                passrate_wo = -1
            passrate = round(summary[key]["passes"]/summary[key]["total"], 3)
            failrate = round(summary[key]["error"]/summary[key]["total"], 3)
            summary_total[key1] = {"passes": summary[key]["passes"], "error": summary[key]["error"], "notsupport": summary[key]["notsupport"], "total": summary[key]["total"], "passrate_wo": passrate_wo, "passrate": passrate, "failrate": failrate}
        else:
            summary_total[key1]["passes"] += summary[key]["passes"]
            summary_total[key1]["error"] += summary[key]["error"]
            summary_total[key1]["notsupport"] += summary[key]["notsupport"]
            summary_total[key1]["total"] += summary[key]["total"]
            if summary[key]["total"] - summary[key]["notsupport"] != 0:
                passrate_wo = round(summary_total[key1]["passes"]/(summary_total[key1]["total"] - summary_total[key1]["notsupport"]), 3)
            else:
                passrate_wo = -1
            passrate = round(summary_total[key1]["passes"]/summary_total[key1]["total"], 3)
            failrate = round(summary_total[key1]["error"]/summary_total[key1]["total"], 3)
            summary_total[key1] = {"passes": summary_total[key1]["passes"], "error": summary_total[key1]["error"], "notsupport": summary_total[key1]["notsupport"], "total": summary_total[key1]["total"], "passrate_wo": passrate_wo, "passrate": passrate, "failrate": failrate}
    elif key == "train" or key == "inf":
        key1 = "total"
        if key1 not in summary_total.keys():
            if summary[key]["total"] - summary[key]["notsupport"] != 0:
                passrate_wo = round(summary[key]["passes"]/(summary[key]["total"] - summary[key]["notsupport"]), 3)
            else:
                passrate_wo = -1
            passrate = round(summary[key]["passes"]/summary[key]["total"], 3)
            failrate = round(summary[key]["error"]/summary[key]["total"], 3)
            summary_total[key1] = {"passes": summary[key]["passes"], "error": summary[key]["error"], "notsupport": summary[key]["notsupport"], "total": summary[key]["total"], "passrate_wo": passrate_wo, "passrate": passrate, "failrate": failrate}
        else:
            summary_total[key1]["passes"] += summary[key]["passes"]
            summary_total[key1]["error"] += summary[key]["error"]
            summary_total[key1]["notsupport"] += summary[key]["notsupport"]
            summary_total[key1]["total"] += summary[key]["total"]
            if summary[key]["total"] - summary[key]["notsupport"] != 0:
                passrate_wo = round(summary_total[key1]["passes"]/(summary_total[key1]["total"] - summary_total[key1]["notsupport"]), 3)
            else:
                passrate_wo = -1
            passrate = round(summary_total[key1]["passes"]/summary_total[key1]["total"], 3)
            failrate = round(summary_total[key1]["error"]/summary_total[key1]["total"], 3)
            summary_total[key1] = {"passes": summary_total[key1]["passes"], "error": summary_total[key1]["error"], "notsupport": summary_total[key1]["notsupport"], "total": summary_total[key1]["total"], "passrate_wo": passrate_wo, "passrate": passrate, "failrate": failrate}


for key in summary_total:
    print((key, {summary_total[key]["passes"]}, {summary_total[key]["error"]}, {summary_total[key]["notsupport"]}, {summary_total[key]["total"]}, summary_total[key]["passrate_wo"], summary_total[key]["passrate"], summary_total[key]["failrate"]))
    ws_s.append((key, summary_total[key]["passes"], summary_total[key]["error"], summary_total[key]["notsupport"], summary_total[key]["total"], summary_total[key]["passrate_wo"], summary_total[key]["passrate"], summary_total[key]["failrate"]))


ws_diff = wb.create_sheet(title="diff")
ws_diff.append(("modelname", "usecase", "backend", "precision", "result","refresult","simple detail","ref_simple", "result details", "ref_details"))
reffile = "refhuggingface.xlsx"
refexcel = load_workbook(reffile, data_only=True)
sheetnames = refexcel.sheetnames
print(sheetnames)


for key in details:
    ws = wb.create_sheet(title=key)
    ws.append(("modelname", "usecase", "backend", "precision", "result", "simple detail", "result details"))
    index = 0
    if "translation" in key or "summarization" in key:
        category = key.split("_")[0]
    elif "image_to_text" in key:
        category = key.split("_")[0] + "_" + key.split("_")[1] + "_"  + key.split("_")[2]
        index = 2
    elif "text_zero_shot_classification" in key:
        category = key.split("_")[0] + "_" + key.split("_")[1] + "_"  + key.split("_")[2] + "_" + key.split("_")[3]
        index = 3
    else:
        category = key.split("_")[0] + "_" + key.split("_")[1]
        index = 1

    usecase = key.split("_")[index + 1]
    #if "optical_flow" in key:
    #    category = "optical_flow"
    #    index = 1
    #else:
    #    category = key.split("_")[0]
    #usecase = key.split("_")[index + 1]
    if "amp_fp16" in key or "amp_bf16" in key:
        precision = key.split("_")[index + 2] + "_" + key.split("_")[index + 3]
        backend = key.split("_")[index + 4]
    else:
        precision = key.split("_")[index + 2]
        backend = key.split("_")[index + 3]
    for result, detail in details[key].items():
        for mymodel in detail:
            result_key = usecase + precision + backend + "_result"
            if result == "error":
                result_detail = results_summary[mymodel][result_key]['failed']
                simple_detail = results_summary[mymodel][result_key]['simple_failed']
                if key in sheetnames:
                    refsheet = refexcel[key]
                    for i in range(2, int(refsheet.max_row)):
                        model_name = refsheet["A" + str(i)].value
                        if model_name == mymodel:
                            ref_result = refsheet["E" + str(i)].value
                            ref_simple = refsheet["F" + str(i)].value if refsheet["F" + str(i)].value else ""
                            ref_detail = refsheet["G" + str(i)].value
                            if ref_result != result or ref_simple != simple_detail:
                                #print(f'{model_name} ref_simple={ref_simple}')
                                #print(f'simple_detail={simple_detail}')
                                ws_diff.append((mymodel, usecase, backend, precision, result,ref_result, simple_detail,ref_simple, result_detail, ref_detail))
            elif result == "notsupport":
                result_detail = results_summary[mymodel][result_key]['Notsupported']
                simple_detail = results_summary[mymodel][result_key]['simple_Notsupported']
                if key in sheetnames:
                    refsheet = refexcel[key]
                    for i in range(2, int(refsheet.max_row)):
                        model_name = refsheet["A" + str(i)].value
                        if model_name == mymodel:
                            ref_result = refsheet["E" + str(i)].value
                            ref_simple = refsheet["F" + str(i)].value if refsheet["F" + str(i)].value else ""
                            ref_detail = refsheet["G" + str(i)].value
                            if ref_result != result or ref_simple != simple_detail:
                                #print(f'{model_name} ref_simple={ref_simple}')
                                #print(f'simple_detail={simple_detail}')
                                ws_diff.append((mymodel, usecase, backend, precision, result,ref_result, simple_detail,ref_simple, result_detail, ref_detail))
            else:
                result_detail = ""
                simple_detail = ""
            ws.append((mymodel, usecase, backend, precision, result, simple_detail, result_detail))
refexcel.close()
wb.save("hugginface.xlsx")



